import os

from openpyxl import Workbook, load_workbook
import pandas as pd



def rewrite_one_wb(wb_path, out_path):
    wb = load_workbook(wb_path)
    ws = wb["Cover Sheet"] 
    # swap loc and fld from HELMS
    if ws['D13'].value.lower()=="roy hill minesite":
        fld_handler = ws["D14"].value
        #loc_handler = "ROY HILL"

        
        ws['D13'].value = fld_handler
        ws['D14'].value = "ROY HILL"
    else: pass
    wb.save(out_path)
    wb.close()

def main():
    current_path = os.path.dirname(__file__)
    load_dir = os.path.join(current_path, r'load_dir')
    out_dir = os.path.join(current_path, r'out_dir')
    las_files = []
    out_paths = []
    # get las files
    for dirpath, dirnames, filenames in os.walk(load_dir):
        for filename in filenames:
            if filename.lower().endswith(".xlsx"):
                las_files.append(os.path.join(dirpath, filename))
                filename1 = filename.replace("RoyHillMinesite", "ROY HILL")
                filename1 = filename1.replace("COMPLETESL", "C")
                out_paths.append(os.path.join(out_dir, filename1))

    for las, out in zip(las_files, out_paths):
        rewrite_one_wb(las, out)


if __name__ == "__main__":
    main()
