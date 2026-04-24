# intended to read well report according to well_report_src.json

from openpyxl import load_workbook
from pathlib import Path
import pandas as pd
import json


class ReadWellReport:
    def __init__(self):
        self.json_load = False
        self.wellreport = None

    def header_info(self, wb_path):
        if not self.json_load:
            raise ValueError("Fetch well report location first")
        field_details = self.wellreport["field_details"]
        wb = load_workbook(wb_path, data_only=True)
        ws = wb.active
        result = {}
        for field_name, cell_ref in field_details.items():
            result[field_name] = ws[cell_ref].value
        return result
    
    def area_info(self, wb_path, area_searched = "logging"):
        if not self.json_load:
            raise ValueError("Fetch well report location first")
        if area_searched != "logging" &  area_searched != "cal":
            raise ValueError("Input correct table area, logging or cal") 
        key = f'{area_searched}_details'
        details = self.wellreport[key]

        wb = load_workbook(wb_path, data_only=True)
        ws = wb.active
        start = details["_table_start_row"]
        end = details["_table_end_row"]

        data = {}

        for k, v in details.items():
            if k.startwith("_"):
                continue
            values = []
            for row in range(start, end+1):
                cell = f"{v}{row}"
                values.append(ws[cell].value)
            
            data[k] = values

        df = pd.DataFrame(data)
        return df # get the final extraction table
 
    def _load_json(self):
        self.json_load = True
        project_root = Path(__file__).resolve().parent
        json_path = project_root / "src" / "config" / "well_report_src.json"
        with open(json_path, "r", encoding="utf-8") as f:
            self.wellreport = json.load(f)
            f.close()



